import json
import logging
import os
import re
import sqlite3
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request

import ccxt
from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, HTTPException, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel

load_dotenv()

# =====================
# CONFIG
# =====================
BINGX_API_KEY = os.getenv("BINGX_API_KEY", "").strip()
BINGX_SECRET = os.getenv("BINGX_SECRET", "").strip()
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "").strip()

ALLOWED_SYMBOLS_RAW = os.getenv("ALLOWED_SYMBOLS", "*").strip()
DEFAULT_SETTLE = os.getenv("DEFAULT_SETTLE", "USDT").strip().upper() or "USDT"
DEFAULT_MARGIN_MODE = os.getenv("DEFAULT_MARGIN_MODE", "cross").strip().lower() or "cross"
PORT = int(os.getenv("PORT", "8000"))
DRY_RUN = os.getenv("DRY_RUN", "true").strip().lower() in {"1", "true", "yes", "on"}

DEDUP_TTL_SEC = int(os.getenv("DEDUP_TTL_SEC", "120"))
MAX_RECENT_EVENTS = int(os.getenv("MAX_RECENT_EVENTS", "2000"))

JOURNAL_DB_PATH = Path(os.getenv("JOURNAL_DB_PATH", "data/hybrid_journal.sqlite3"))
REPORTS_DIR = Path(os.getenv("REPORTS_DIR", "reports"))
REPORT_FILENAME = os.getenv("REPORT_FILENAME", "bingx_hybrid_report.xlsx").strip() or "bingx_hybrid_report.xlsx"
RECONCILE_POLL_COUNT = int(os.getenv("RECONCILE_POLL_COUNT", "3"))
RECONCILE_POLL_DELAY_SEC = float(os.getenv("RECONCILE_POLL_DELAY_SEC", "0.8"))
LOG_TG_TOKEN = os.getenv("LOG_TG_TOKEN", "").strip()
LOG_TG_CHAT_ID = os.getenv("LOG_TG_CHAT_ID", "").strip()
LOG_TG_TIMEOUT_SEC = float(os.getenv("LOG_TG_TIMEOUT_SEC", "10"))

if ALLOWED_SYMBOLS_RAW == "*":
    ALLOWED_SYMBOLS = {"*"}
else:
    ALLOWED_SYMBOLS = {s.strip().upper().replace("-", "") for s in ALLOWED_SYMBOLS_RAW.split(",") if s.strip()}

# =====================
# LOGGING
# =====================
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger("bingx_hybrid_bot")

# =====================
# APP
# =====================
app = FastAPI(title="BingX Hybrid Webhook Bot + Reporting")

exchange = ccxt.bingx(
    {
        "apiKey": BINGX_API_KEY,
        "secret": BINGX_SECRET,
        "enableRateLimit": True,
        "options": {
            "defaultType": "swap",
            "defaultMarginMode": DEFAULT_MARGIN_MODE,
        },
    }
)

_markets_loaded = False
_markets_lock = threading.Lock()
_recent_events: Dict[str, float] = {}
_recent_lock = threading.Lock()
_log_tg_lock = threading.Lock()
_cached_log_tg_chat_id: Optional[str] = LOG_TG_CHAT_ID or None


class WebhookPayload(BaseModel):
    source: Optional[str] = None
    mode: Optional[str] = None
    token: Optional[str] = None

    action: str
    symbol: str
    side: Optional[str] = None
    reason: Optional[str] = None

    orderId: Optional[str] = None
    lotTag: Optional[str] = None

    qtyCoin: Optional[float] = None
    triggerPrice: Optional[float] = None
    fillPrice: Optional[float] = None
    entryPrice: Optional[float] = None
    theoreticalAvg: Optional[float] = None
    theoreticalPositionQty: Optional[float] = None

    tpPercent: Optional[float] = None
    tpPrice: Optional[float] = None
    lotTpPrice: Optional[float] = None
    currentClose: Optional[float] = None
    callbackPercent: Optional[float] = None
    trailingMin: Optional[float] = None
    trailStop: Optional[float] = None

    numSells: Optional[int] = None
    subCoverMode: Optional[str] = None

    barTime: Optional[int] = None
    timestamp: Optional[int] = None


def now_iso() -> str:
    return datetime.utcnow().isoformat() + "Z"


def has_exchange_credentials() -> bool:
    return bool(BINGX_API_KEY and BINGX_SECRET)


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def optional_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def safe_json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, sort_keys=True, default=str)


def new_event_id() -> str:
    return uuid.uuid4().hex


def ensure_storage_ready() -> None:
    JOURNAL_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def first_number(*values: Any, allow_zero: bool = False) -> Optional[float]:
    for value in values:
        number = optional_float(value)
        if number is None:
            continue
        if allow_zero or number != 0:
            return number
    return None


def first_value(*values: Any) -> Any:
    for value in values:
        if value is None or value == "":
            continue
        return value
    return None


def format_number(value: Any, *, digits: int = 8) -> str:
    number = optional_float(value)
    if number is None:
        return "-"
    text = f"{number:.{digits}f}".rstrip("0").rstrip(".")
    return text if text else "0"


def format_value(value: Any) -> str:
    if value is None or value == "":
        return "-"
    return str(value)


def snapshot_metric(snapshot: Optional[Dict[str, Any]], *keys: str, allow_zero: bool = True) -> Optional[float]:
    if not snapshot:
        return None
    return first_number(*(snapshot.get(key) for key in keys), allow_zero=allow_zero)


def snapshot_value(snapshot: Optional[Dict[str, Any]], *keys: str) -> Any:
    if not snapshot:
        return None
    return first_value(*(snapshot.get(key) for key in keys))


def snapshot_to_report_dict(snapshot: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    snap = snapshot or {}
    return {
        "side": snapshot_value(snap, "side"),
        "contracts": snapshot_metric(snap, "contracts", allow_zero=True),
        "entry_price": snapshot_metric(snap, "entry_price", "entryPrice"),
        "mark_price": snapshot_metric(snap, "mark_price", "markPrice"),
        "unrealized_pnl": snapshot_metric(snap, "unrealized_pnl", "unrealizedPnl", allow_zero=True),
        "market_price": snapshot_metric(snap, "market_price"),
        "order_id": snapshot_value(snap, "order_id", "orderId"),
        "extra_json": _load_json_dict(snapshot_value(snap, "extra_json", "extra")),
    }


def extract_order_execution_price(order: Optional[Dict[str, Any]], fallback_price: Any = None) -> Optional[float]:
    info = order.get("info") if isinstance(order, dict) and isinstance(order.get("info"), dict) else {}
    return first_number(
        order.get("average") if isinstance(order, dict) else None,
        order.get("price") if isinstance(order, dict) else None,
        info.get("avgPrice"),
        info.get("averagePrice"),
        info.get("price"),
        info.get("dealPrice"),
        info.get("executedPrice"),
        fallback_price,
    )


def telegram_logging_enabled() -> bool:
    return bool(LOG_TG_TOKEN)


def telegram_api_call(method: str, params: Dict[str, Any]) -> Any:
    encoded = urllib_parse.urlencode({k: v for k, v in params.items() if v is not None}).encode("utf-8")
    request = urllib_request.Request(
        f"https://api.telegram.org/bot{LOG_TG_TOKEN}/{method}",
        data=encoded,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    with urllib_request.urlopen(request, timeout=max(LOG_TG_TIMEOUT_SEC, 1.0)) as response:
        payload = json.loads(response.read().decode("utf-8"))
    if not payload.get("ok"):
        raise RuntimeError(payload.get("description") or f"Telegram API {method} failed")
    return payload.get("result")


def extract_chat_id_from_update(update: Dict[str, Any]) -> Optional[str]:
    chat_sources = [
        update.get("message"),
        update.get("edited_message"),
        update.get("channel_post"),
        update.get("edited_channel_post"),
        (update.get("callback_query") or {}).get("message"),
        update.get("my_chat_member"),
        update.get("chat_member"),
    ]
    for source in chat_sources:
        if not isinstance(source, dict):
            continue
        chat = source.get("chat")
        if isinstance(chat, dict) and chat.get("id") is not None:
            return str(chat["id"])
    return None


def resolve_log_tg_chat_id() -> Optional[str]:
    global _cached_log_tg_chat_id
    if _cached_log_tg_chat_id:
        return _cached_log_tg_chat_id
    if not telegram_logging_enabled():
        return None

    with _log_tg_lock:
        if _cached_log_tg_chat_id:
            return _cached_log_tg_chat_id
        try:
            updates = telegram_api_call("getUpdates", {"limit": 20, "timeout": 0}) or []
        except Exception as exc:
            logger.warning("Telegram chat_id auto-discovery failed: %s", exc)
            return None

        for update in reversed(updates):
            chat_id = extract_chat_id_from_update(update if isinstance(update, dict) else {})
            if chat_id:
                _cached_log_tg_chat_id = chat_id
                logger.info("Resolved Telegram log chat_id automatically: %s", chat_id)
                return chat_id

    return None


def chunk_telegram_message(text: str, limit: int = 3900) -> List[str]:
    if len(text) <= limit:
        return [text]
    chunks: List[str] = []
    current = ""
    for line in text.splitlines():
        candidate = line if not current else f"{current}\n{line}"
        if len(candidate) <= limit:
            current = candidate
            continue
        if current:
            chunks.append(current)
        current = line
    if current:
        chunks.append(current)
    return chunks or [text[:limit]]


def send_telegram_log(text: str) -> None:
    if not telegram_logging_enabled():
        return
    chat_id = resolve_log_tg_chat_id()
    if not chat_id:
        logger.warning("Telegram logging skipped: LOG_TG_CHAT_ID is not set and no chat_id could be auto-discovered")
        return
    for chunk in chunk_telegram_message(text):
        try:
            telegram_api_call("sendMessage", {"chat_id": chat_id, "text": chunk})
        except (urllib_error.URLError, RuntimeError, ValueError) as exc:
            logger.warning("Telegram log send failed: %s", exc)
            return


def build_alert_accept_message(event_id: str, payload: WebhookPayload) -> str:
    lines = [
        "ALERT ACCEPTED",
        f"event_id={event_id}",
        f"action={format_value(payload.action)}",
        f"symbol={format_value(payload.symbol)}",
        f"dry_run={DRY_RUN}",
        f"order_id={format_value(payload.orderId)}",
        f"lot_tag={format_value(payload.lotTag)}",
        f"reason={format_value(payload.reason)}",
        f"qty_coin={format_number(payload.qtyCoin)}",
        f"trigger_price={format_number(payload.triggerPrice)}",
        f"fill_price={format_number(payload.fillPrice)}",
        f"entry_price={format_number(payload.entryPrice)}",
        f"theoretical_avg={format_number(payload.theoreticalAvg)}",
        f"tp_price={format_number(payload.tpPrice)}",
        f"lot_tp_price={format_number(payload.lotTpPrice)}",
        f"current_close={format_number(payload.currentClose)}",
        f"timestamp={format_value(payload.timestamp)}",
    ]
    return "\n".join(lines)


def build_duplicate_alert_message(payload: WebhookPayload) -> str:
    lines = [
        "ALERT DUPLICATE IGNORED",
        f"action={format_value(payload.action)}",
        f"symbol={format_value(payload.symbol)}",
        f"order_id={format_value(payload.orderId)}",
        f"lot_tag={format_value(payload.lotTag)}",
        f"qty_coin={format_number(payload.qtyCoin)}",
        f"timestamp={format_value(payload.timestamp)}",
    ]
    return "\n".join(lines)


def build_result_log_message(event_id: str, payload: WebhookPayload, result: Dict[str, Any]) -> str:
    reconcile = result.get("reconcile", {}) if isinstance(result.get("reconcile"), dict) else {}
    pre_snapshot = snapshot_to_report_dict(reconcile.get("pre_snapshot"))
    post_snapshot = snapshot_to_report_dict(reconcile.get("post_snapshot"))

    action = str(payload.action or "").upper()
    if action in {"FIRST_SHORT", "RESTART_SHORT", "DCA_SHORT"}:
        actual_entry_price = first_number(
            result.get("entry_price"),
            post_snapshot.get("entry_price"),
            payload.fillPrice,
            payload.entryPrice,
            payload.theoreticalAvg,
        )
        actual_exit_price = None
    else:
        actual_entry_price = first_number(
            result.get("entry_price"),
            pre_snapshot.get("entry_price"),
            payload.entryPrice,
            payload.theoreticalAvg,
        )
        actual_exit_price = first_number(
            result.get("exit_price"),
            result.get("market_price"),
            payload.currentClose,
            payload.fillPrice,
        )

    lines = [
        "ALERT RESULT",
        f"event_id={event_id}",
        f"action={format_value(payload.action)}",
        f"symbol={format_value(result.get('symbol') or payload.symbol)}",
        f"status={format_value(result.get('status'))}",
        f"dry_run={DRY_RUN}",
        f"order_id={format_value(result.get('order_id') or payload.orderId)}",
        f"reconcile_status={format_value(reconcile.get('reconcile_status'))}",
        f"requested_qty={format_number(first_number(payload.qtyCoin, result.get('qty'), result.get('requested_qty'), allow_zero=True))}",
        f"closed_qty={format_number(first_number(result.get('closed_qty'), allow_zero=True))}",
        f"actual_entry_price={format_number(actual_entry_price)}",
        f"actual_exit_price={format_number(actual_exit_price)}",
        f"market_price={format_number(first_number(result.get('market_price'), post_snapshot.get('market_price')))}",
        f"tp_price={format_number(first_number(result.get('tp_price'), payload.tpPrice))}",
        f"lot_tp_price={format_number(first_number(result.get('lot_tp_price'), payload.lotTpPrice))}",
        f"pre_contracts={format_number(pre_snapshot.get('contracts'))}",
        f"pre_entry_price={format_number(pre_snapshot.get('entry_price'))}",
        f"post_contracts={format_number(post_snapshot.get('contracts'))}",
        f"post_entry_price={format_number(post_snapshot.get('entry_price'))}",
    ]

    error_text = result.get("error_text")
    if error_text:
        lines.append(f"error={format_value(error_text)}")

    return "\n".join(lines)


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(JOURNAL_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    ensure_storage_ready()
    with get_db_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS events (
                event_id TEXT PRIMARY KEY,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                action TEXT NOT NULL,
                status TEXT NOT NULL,
                source TEXT,
                mode TEXT,
                dry_run INTEGER NOT NULL,
                symbol_raw TEXT,
                symbol_compact TEXT,
                ccxt_symbol TEXT,
                order_id TEXT,
                lot_tag TEXT,
                reason TEXT,
                error_text TEXT,
                request_json TEXT NOT NULL,
                result_json TEXT
            );

            CREATE TABLE IF NOT EXISTS position_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id TEXT NOT NULL,
                created_at TEXT NOT NULL,
                stage TEXT NOT NULL,
                symbol_compact TEXT,
                ccxt_symbol TEXT,
                side TEXT,
                contracts REAL,
                entry_price REAL,
                mark_price REAL,
                unrealized_pnl REAL,
                market_price REAL,
                order_id TEXT,
                extra_json TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_events_created_at ON events(created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_events_action ON events(action);
            CREATE INDEX IF NOT EXISTS idx_events_status ON events(status);
            CREATE INDEX IF NOT EXISTS idx_snapshots_event_id ON position_snapshots(event_id);
            CREATE INDEX IF NOT EXISTS idx_snapshots_created_at ON position_snapshots(created_at DESC);
            """
        )


def insert_event_shell(event_id: str, payload: WebhookPayload) -> None:
    now = now_iso()
    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO events (
                event_id, created_at, updated_at, action, status, source, mode, dry_run,
                symbol_raw, lot_tag, reason, request_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_id,
                now,
                now,
                payload.action,
                "accepted",
                payload.source,
                payload.mode,
                int(DRY_RUN),
                payload.symbol,
                payload.lotTag,
                payload.reason,
                safe_json(payload.model_dump()),
            ),
        )


def update_event_status(
    event_id: str,
    *,
    status: str,
    compact_symbol: Optional[str] = None,
    ccxt_symbol: Optional[str] = None,
    order_id: Optional[str] = None,
    result: Optional[Dict[str, Any]] = None,
    error_text: Optional[str] = None,
) -> None:
    with get_db_connection() as conn:
        conn.execute(
            """
            UPDATE events
               SET updated_at = ?,
                   status = ?,
                   symbol_compact = COALESCE(?, symbol_compact),
                   ccxt_symbol = COALESCE(?, ccxt_symbol),
                   order_id = COALESCE(?, order_id),
                   result_json = COALESCE(?, result_json),
                   error_text = COALESCE(?, error_text)
             WHERE event_id = ?
            """,
            (
                now_iso(),
                status,
                compact_symbol,
                ccxt_symbol,
                order_id,
                safe_json(result) if result is not None else None,
                error_text,
                event_id,
            ),
        )


def record_snapshot(
    event_id: str,
    stage: str,
    compact_symbol: str,
    ccxt_symbol: str,
    snapshot: Optional[Dict[str, Any]],
    *,
    market_price: Optional[float] = None,
    order_id: Optional[str] = None,
    extra: Optional[Dict[str, Any]] = None,
) -> None:
    snap = snapshot or {}
    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO position_snapshots (
                event_id, created_at, stage, symbol_compact, ccxt_symbol, side,
                contracts, entry_price, mark_price, unrealized_pnl, market_price,
                order_id, extra_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_id,
                now_iso(),
                stage,
                compact_symbol,
                ccxt_symbol,
                snap.get("side"),
                safe_float(snap.get("contracts"), 0.0),
                safe_float(snap.get("entryPrice"), 0.0),
                safe_float(snap.get("markPrice"), 0.0),
                safe_float(snap.get("unrealizedPnl"), 0.0),
                market_price,
                order_id,
                safe_json(extra or {}),
            ),
        )


def ensure_exchange_ready() -> None:
    global _markets_loaded
    if not has_exchange_credentials():
        raise RuntimeError("BINGX_API_KEY/BINGX_SECRET are not configured")
    if _markets_loaded:
        return
    with _markets_lock:
        if _markets_loaded:
            return
        exchange.load_markets()
        _markets_loaded = True
        logger.info("Markets loaded: %s", len(exchange.markets))


def normalize_tv_symbol_to_ccxt(raw_symbol: str) -> Tuple[str, str]:
    s = (raw_symbol or "").strip().upper()
    if not s:
        raise ValueError("Empty symbol")

    if ":" in s and "/" not in s:
        s = s.split(":", 1)[1]

    s = re.sub(r"(\.P|[-_]?PERP|[-_]?SWAP)$", "", s)

    if "/" in s and ":" in s:
        compact = s.split(":", 1)[0].replace("/", "")
        return s, compact

    if "/" in s and ":" not in s:
        base, quote = s.split("/", 1)
        base = re.sub(r"[^A-Z0-9]", "", base)
        quote = re.sub(r"[^A-Z0-9]", "", quote)
        return f"{base}/{quote}:{quote}", f"{base}{quote}"

    compact = re.sub(r"[^A-Z0-9]", "", s)
    compact = re.sub(r"(PERP|SWAP|P)$", "", compact)

    if compact.endswith(DEFAULT_SETTLE):
        base = compact[: -len(DEFAULT_SETTLE)]
        quote = DEFAULT_SETTLE
        if not base:
            raise ValueError(f"Invalid symbol: {raw_symbol}")
        return f"{base}/{quote}:{quote}", compact

    raise ValueError(f"Cannot normalize symbol '{raw_symbol}'. Expected BTCUSDT or BTC/USDT:USDT")


def is_symbol_allowed(compact_symbol: str) -> bool:
    return "*" in ALLOWED_SYMBOLS or compact_symbol.upper() in ALLOWED_SYMBOLS


def verify_secret(payload: WebhookPayload, request: Request) -> None:
    if not WEBHOOK_SECRET:
        return

    provided = payload.token or request.headers.get("X-Webhook-Secret") or request.query_params.get("token")
    if provided != WEBHOOK_SECRET:
        raise HTTPException(status_code=401, detail="Invalid webhook secret")


def cleanup_recent_events() -> None:
    cutoff = time.time() - DEDUP_TTL_SEC
    stale_keys = [k for k, ts in _recent_events.items() if ts < cutoff]
    for key in stale_keys:
        _recent_events.pop(key, None)

    if len(_recent_events) > MAX_RECENT_EVENTS:
        overflow = len(_recent_events) - MAX_RECENT_EVENTS
        for key in sorted(_recent_events, key=_recent_events.get)[:overflow]:
            _recent_events.pop(key, None)


def make_dedup_key(payload: WebhookPayload) -> str:
    return "|".join(
        [
            payload.action or "",
            payload.symbol or "",
            payload.orderId or "",
            payload.lotTag or "",
            str(payload.timestamp or ""),
            str(payload.qtyCoin or ""),
        ]
    )


def register_event_once(payload: WebhookPayload) -> bool:
    key = make_dedup_key(payload)
    now_ts = time.time()
    with _recent_lock:
        cleanup_recent_events()
        if key in _recent_events:
            return False
        _recent_events[key] = now_ts
        return True


def fetch_position_snapshot(ccxt_symbol: str) -> Dict[str, Any]:
    if not has_exchange_credentials():
        return {
            "symbol": ccxt_symbol,
            "side": "unknown",
            "contracts": 0.0,
            "entryPrice": 0.0,
            "markPrice": 0.0,
            "unrealizedPnl": 0.0,
            "raw": None,
            "exchange_state_available": False,
        }

    ensure_exchange_ready()
    positions = exchange.fetch_positions([ccxt_symbol])
    for pos in positions:
        if pos.get("symbol") != ccxt_symbol:
            continue
        side = str(pos.get("side") or "").lower() or "unknown"
        contracts = abs(safe_float(pos.get("contracts") or pos.get("contractSize"), 0.0))
        return {
            "symbol": ccxt_symbol,
            "side": side,
            "contracts": contracts,
            "entryPrice": safe_float(pos.get("entryPrice"), 0.0),
            "markPrice": safe_float(pos.get("markPrice"), 0.0),
            "unrealizedPnl": safe_float(pos.get("unrealizedPnl"), 0.0),
            "raw": pos,
            "exchange_state_available": True,
        }

    return {
        "symbol": ccxt_symbol,
        "side": "flat",
        "contracts": 0.0,
        "entryPrice": 0.0,
        "markPrice": 0.0,
        "unrealizedPnl": 0.0,
        "raw": None,
        "exchange_state_available": True,
    }


def get_real_short_position(ccxt_symbol: str) -> Optional[Dict[str, Any]]:
    snapshot = fetch_position_snapshot(ccxt_symbol)
    if snapshot.get("side") == "short" and safe_float(snapshot.get("contracts"), 0.0) > 0:
        return snapshot
    return None


def fetch_last_price(ccxt_symbol: str) -> float:
    ensure_exchange_ready()
    ticker = exchange.fetch_ticker(ccxt_symbol)
    last = ticker.get("last") or ticker.get("close") or ticker.get("mark")
    if last is None:
        raise RuntimeError(f"No market price for {ccxt_symbol}")
    return float(last)


def place_short_market_entry(ccxt_symbol: str, qty_coin: float, payload: WebhookPayload) -> Dict[str, Any]:
    if qty_coin <= 0:
        raise ValueError("qtyCoin must be > 0 for entry")

    if DRY_RUN:
        logger.info("[DRY_RUN] SELL market %s qty=%s action=%s", ccxt_symbol, qty_coin, payload.action)
        return {"id": "dry-run-entry", "symbol": ccxt_symbol, "amount": qty_coin, "side": "sell", "dry_run": True}

    order = exchange.create_order(symbol=ccxt_symbol, type="market", side="sell", amount=qty_coin, params={})
    logger.info("ENTRY executed | action=%s | symbol=%s | qty=%s | order_id=%s", payload.action, ccxt_symbol, qty_coin, order.get("id"))
    return order


def close_short_reduce_only(ccxt_symbol: str, qty_coin: float, reason: str) -> Dict[str, Any]:
    if qty_coin <= 0:
        raise ValueError("close qty must be > 0")

    params = {"reduceOnly": True}

    if DRY_RUN:
        logger.info("[DRY_RUN] BUY reduceOnly %s qty=%s reason=%s", ccxt_symbol, qty_coin, reason)
        return {
            "id": "dry-run-close",
            "symbol": ccxt_symbol,
            "amount": qty_coin,
            "side": "buy",
            "reduceOnly": True,
            "dry_run": True,
        }

    order = exchange.create_order(symbol=ccxt_symbol, type="market", side="buy", amount=qty_coin, params=params)
    logger.info("CLOSE executed | reason=%s | symbol=%s | qty=%s | order_id=%s", reason, ccxt_symbol, qty_coin, order.get("id"))
    return order


def reconcile_after_order(ccxt_symbol: str, pre_snapshot: Dict[str, Any]) -> Dict[str, Any]:
    if DRY_RUN or not has_exchange_credentials():
        pre_contracts = safe_float(pre_snapshot.get("contracts"), 0.0)
        return {
            "reconcile_status": "dry_run" if DRY_RUN else "no_exchange_credentials",
            "pre_snapshot": pre_snapshot,
            "pre_contracts": pre_contracts,
            "post_contracts": pre_contracts,
            "post_snapshot": pre_snapshot,
        }

    pre_contracts = safe_float(pre_snapshot.get("contracts"), 0.0)
    post_snapshot = pre_snapshot
    for attempt in range(max(RECONCILE_POLL_COUNT, 1)):
        post_snapshot = fetch_position_snapshot(ccxt_symbol)
        post_contracts = safe_float(post_snapshot.get("contracts"), 0.0)
        if post_contracts != pre_contracts or post_contracts == 0 or attempt == RECONCILE_POLL_COUNT - 1:
            break
        time.sleep(max(RECONCILE_POLL_DELAY_SEC, 0.0))

    post_contracts = safe_float(post_snapshot.get("contracts"), 0.0)
    if pre_contracts == 0 and post_contracts > 0:
        reconcile_status = "opened"
    elif pre_contracts > 0 and post_contracts == 0:
        reconcile_status = "fully_closed"
    elif post_contracts < pre_contracts:
        reconcile_status = "reduced"
    elif post_contracts > pre_contracts:
        reconcile_status = "increased"
    else:
        reconcile_status = "unchanged"

    return {
        "reconcile_status": reconcile_status,
        "pre_snapshot": pre_snapshot,
        "pre_contracts": pre_contracts,
        "post_contracts": post_contracts,
        "post_snapshot": post_snapshot,
    }


def handle_entry_like_action(payload: WebhookPayload, event_id: str, ccxt_symbol: str, compact_symbol: str) -> Dict[str, Any]:
    qty = safe_float(payload.qtyCoin, 0.0)
    pre_snapshot = fetch_position_snapshot(ccxt_symbol)
    record_snapshot(event_id, "pre_entry", compact_symbol, ccxt_symbol, pre_snapshot)

    order = place_short_market_entry(ccxt_symbol, qty, payload)
    reconcile = reconcile_after_order(ccxt_symbol, pre_snapshot)
    entry_price = first_number(
        extract_order_execution_price(order, payload.fillPrice),
        snapshot_metric(reconcile.get("post_snapshot"), "entryPrice"),
        payload.fillPrice,
        payload.entryPrice,
        payload.theoreticalAvg,
    )
    record_snapshot(
        event_id,
        "post_entry",
        compact_symbol,
        ccxt_symbol,
        reconcile["post_snapshot"],
        order_id=order.get("id"),
        extra={"reconcile_status": reconcile["reconcile_status"], "entry_price": entry_price},
    )

    return {
        "status": "executed",
        "action": payload.action,
        "symbol": ccxt_symbol,
        "compact_symbol": compact_symbol,
        "qty": qty,
        "entry_price": entry_price,
        "order_id": order.get("id"),
        "reconcile": reconcile,
    }


def handle_full_tp_close(payload: WebhookPayload, event_id: str, ccxt_symbol: str, compact_symbol: str) -> Dict[str, Any]:
    real_pos = get_real_short_position(ccxt_symbol)
    record_snapshot(event_id, "pre_full_tp", compact_symbol, ccxt_symbol, real_pos)

    if not real_pos:
        logger.info("No short position for %s; FULL_TP_CLOSE ignored", ccxt_symbol)
        return {"status": "no_position", "action": payload.action, "symbol": ccxt_symbol, "compact_symbol": compact_symbol}

    real_avg = safe_float(real_pos["entryPrice"], 0.0)
    qty = safe_float(real_pos["contracts"], 0.0)
    tp_percent = safe_float(payload.tpPercent, 0.0)
    if tp_percent <= 0:
        raise ValueError("tpPercent must be > 0 for FULL_TP_CLOSE")

    tp_price = real_avg * (1.0 - tp_percent / 100.0)
    current_price = fetch_last_price(ccxt_symbol)

    logger.info(
        "FULL_TP check | symbol=%s | real_avg=%.8f | tp_percent=%.4f | tp_price=%.8f | market=%.8f | qty=%.8f",
        ccxt_symbol,
        real_avg,
        tp_percent,
        tp_price,
        current_price,
        qty,
    )

    if current_price > tp_price:
        logger.warning("FULL_TP condition not met | symbol=%s | market=%.8f | tp=%.8f", ccxt_symbol, current_price, tp_price)
        record_snapshot(
            event_id,
            "tp_rejected",
            compact_symbol,
            ccxt_symbol,
            real_pos,
            market_price=current_price,
            extra={"tp_price": tp_price, "reason": "tp_not_reached"},
        )
        return {
            "status": "tp_not_reached",
            "action": payload.action,
            "symbol": ccxt_symbol,
            "compact_symbol": compact_symbol,
            "real_avg": real_avg,
            "tp_price": tp_price,
            "market_price": current_price,
        }

    order = close_short_reduce_only(ccxt_symbol, qty, payload.reason or "FULL_TP_CLOSE")
    reconcile = reconcile_after_order(ccxt_symbol, real_pos)
    exit_price = first_number(extract_order_execution_price(order, current_price), current_price, payload.currentClose)
    record_snapshot(
        event_id,
        "post_full_tp",
        compact_symbol,
        ccxt_symbol,
        reconcile["post_snapshot"],
        market_price=current_price,
        order_id=order.get("id"),
        extra={"tp_price": tp_price, "reconcile_status": reconcile["reconcile_status"], "exit_price": exit_price, "entry_price": real_avg, "closed_qty": qty},
    )
    return {
        "status": "executed",
        "action": payload.action,
        "symbol": ccxt_symbol,
        "compact_symbol": compact_symbol,
        "entry_price": real_avg,
        "exit_price": exit_price,
        "real_avg": real_avg,
        "tp_price": tp_price,
        "market_price": current_price,
        "closed_qty": qty,
        "order_id": order.get("id"),
        "reconcile": reconcile,
    }


def handle_subcover_close(payload: WebhookPayload, event_id: str, ccxt_symbol: str, compact_symbol: str) -> Dict[str, Any]:
    real_pos = get_real_short_position(ccxt_symbol)
    record_snapshot(event_id, "pre_subcover", compact_symbol, ccxt_symbol, real_pos)

    if not real_pos:
        logger.info("No short position for %s; SUBCOVER_CLOSE ignored", ccxt_symbol)
        return {"status": "no_position", "action": payload.action, "symbol": ccxt_symbol, "compact_symbol": compact_symbol}

    requested_qty = safe_float(payload.qtyCoin, 0.0)
    if requested_qty <= 0:
        raise ValueError("qtyCoin must be > 0 for SUBCOVER_CLOSE")

    real_qty = safe_float(real_pos["contracts"], 0.0)
    close_qty = min(requested_qty, real_qty)
    current_price = fetch_last_price(ccxt_symbol)

    lot_tp = safe_float(payload.lotTpPrice, 0.0)
    if lot_tp > 0 and current_price > lot_tp:
        logger.warning("SUBCOVER condition not met | symbol=%s | market=%.8f | lot_tp=%.8f", ccxt_symbol, current_price, lot_tp)
        record_snapshot(
            event_id,
            "subcover_rejected",
            compact_symbol,
            ccxt_symbol,
            real_pos,
            market_price=current_price,
            extra={"lot_tp_price": lot_tp, "reason": "tp_not_reached"},
        )
        return {
            "status": "tp_not_reached",
            "action": payload.action,
            "symbol": ccxt_symbol,
            "compact_symbol": compact_symbol,
            "lot_tp_price": lot_tp,
            "market_price": current_price,
        }

    order = close_short_reduce_only(ccxt_symbol, close_qty, "SUBCOVER_CLOSE")
    reconcile = reconcile_after_order(ccxt_symbol, real_pos)
    entry_price = first_number(payload.entryPrice, safe_float(real_pos.get("entryPrice"), 0.0))
    exit_price = first_number(extract_order_execution_price(order, current_price), current_price, payload.currentClose)
    record_snapshot(
        event_id,
        "post_subcover",
        compact_symbol,
        ccxt_symbol,
        reconcile["post_snapshot"],
        market_price=current_price,
        order_id=order.get("id"),
        extra={
            "requested_qty": requested_qty,
            "closed_qty": close_qty,
            "reconcile_status": reconcile["reconcile_status"],
            "entry_price": entry_price,
            "exit_price": exit_price,
            "lot_tp_price": lot_tp,
        },
    )
    return {
        "status": "executed",
        "action": payload.action,
        "symbol": ccxt_symbol,
        "compact_symbol": compact_symbol,
        "entry_price": entry_price,
        "exit_price": exit_price,
        "lot_tp_price": lot_tp,
        "requested_qty": requested_qty,
        "closed_qty": close_qty,
        "market_price": current_price,
        "order_id": order.get("id"),
        "reconcile": reconcile,
    }


def process_payload(payload: WebhookPayload, event_id: str) -> Dict[str, Any]:
    ccxt_symbol, compact_symbol = normalize_tv_symbol_to_ccxt(payload.symbol)
    if not is_symbol_allowed(compact_symbol):
        logger.warning("Symbol not allowed: %s (normalized %s)", payload.symbol, compact_symbol)
        return {"status": "ignored", "reason": "symbol_not_allowed", "symbol": compact_symbol, "compact_symbol": compact_symbol}

    action = payload.action.strip().upper()

    if action == "TEST_ALERT":
        snapshot = fetch_position_snapshot(ccxt_symbol)
        record_snapshot(event_id, "test_alert", compact_symbol, ccxt_symbol, snapshot, extra={"reason": payload.reason or "MANUAL_HYBRID_DEBUG"})
        return {
            "status": "test_ok",
            "action": action,
            "symbol": ccxt_symbol,
            "compact_symbol": compact_symbol,
            "reason": payload.reason or "MANUAL_HYBRID_DEBUG",
        }

    if action in {"FIRST_SHORT", "RESTART_SHORT", "DCA_SHORT"}:
        return handle_entry_like_action(payload, event_id, ccxt_symbol, compact_symbol)

    if action == "FULL_TP_CLOSE":
        return handle_full_tp_close(payload, event_id, ccxt_symbol, compact_symbol)

    if action == "SUBCOVER_CLOSE":
        return handle_subcover_close(payload, event_id, ccxt_symbol, compact_symbol)

    logger.info("Unsupported action ignored: %s", action)
    return {"status": "ignored", "reason": "unsupported_action", "action": action, "compact_symbol": compact_symbol, "symbol": ccxt_symbol}


def background_process(payload: WebhookPayload, event_id: str) -> None:
    try:
        result = process_payload(payload, event_id)
        update_event_status(
            event_id,
            status=result.get("status", "processed"),
            compact_symbol=result.get("compact_symbol"),
            ccxt_symbol=result.get("symbol"),
            order_id=result.get("order_id"),
            result=result,
        )
        logger.info("Webhook processed | event_id=%s | action=%s | result=%s", event_id, payload.action, result)
        send_telegram_log(build_result_log_message(event_id, payload, result))
    except Exception as exc:
        update_event_status(event_id, status="error", error_text=str(exc))
        logger.exception("Webhook processing failed | event_id=%s | action=%s | symbol=%s | error=%s", event_id, payload.action, payload.symbol, exc)
        send_telegram_log(
            "\n".join(
                [
                    "ALERT ERROR",
                    f"event_id={event_id}",
                    f"action={format_value(payload.action)}",
                    f"symbol={format_value(payload.symbol)}",
                    f"error={format_value(exc)}",
                ]
            )
        )


def run_startup_sync() -> None:
    event_id = new_event_id()
    payload = WebhookPayload(action="STARTUP_SYNC", symbol="SYSTEM")
    insert_event_shell(event_id, payload)

    if not has_exchange_credentials():
        result = {"status": "startup_skipped", "reason": "no_exchange_credentials"}
        update_event_status(event_id, status=result["status"], result=result)
        return

    synced = 0
    details: List[Dict[str, Any]] = []
    try:
        if "*" in ALLOWED_SYMBOLS:
            result = {
                "status": "startup_partial",
                "reason": "ALLOWED_SYMBOLS=*; explicit snapshot sync skipped to avoid full account scan",
            }
            update_event_status(event_id, status=result["status"], result=result)
            return

        for compact_symbol in sorted(ALLOWED_SYMBOLS):
            ccxt_symbol, normalized_compact = normalize_tv_symbol_to_ccxt(compact_symbol)
            snapshot = fetch_position_snapshot(ccxt_symbol)
            record_snapshot(event_id, "startup_sync", normalized_compact, ccxt_symbol, snapshot)
            synced += 1
            details.append({"compact_symbol": normalized_compact, "contracts": snapshot.get("contracts", 0.0), "side": snapshot.get("side")})

        result = {"status": "startup_synced", "symbols_synced": synced, "details": details}
        update_event_status(event_id, status=result["status"], result=result)
    except Exception as exc:
        update_event_status(event_id, status="startup_error", error_text=str(exc))
        logger.exception("Startup sync failed: %s", exc)


@app.on_event("startup")
def on_startup() -> None:
    init_db()
    if not has_exchange_credentials():
        logger.warning("BINGX_API_KEY or BINGX_SECRET are not configured. Exchange calls will be skipped.")
        run_startup_sync()
        return
    try:
        ensure_exchange_ready()
    except Exception as exc:
        logger.exception("Failed to load markets on startup: %s", exc)
    run_startup_sync()


@app.on_event("shutdown")
def on_shutdown() -> None:
    try:
        exchange.close()
    except Exception:
        pass


def fetch_recent_events(limit: int = 200) -> List[Dict[str, Any]]:
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT event_id, created_at, updated_at, action, status, source, mode, dry_run,
                   symbol_raw, symbol_compact, ccxt_symbol, order_id, lot_tag, reason,
                   error_text, request_json, result_json
              FROM events
          ORDER BY created_at DESC
             LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def fetch_recent_snapshots(limit: int = 500) -> List[Dict[str, Any]]:
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, event_id, created_at, stage, symbol_compact, ccxt_symbol, side,
                   contracts, entry_price, mark_price, unrealized_pnl, market_price,
                   order_id, extra_json
              FROM position_snapshots
          ORDER BY created_at DESC, id DESC
             LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def build_summary() -> Dict[str, Any]:
    with get_db_connection() as conn:
        total_events = conn.execute("SELECT COUNT(*) FROM events").fetchone()[0]
        total_snapshots = conn.execute("SELECT COUNT(*) FROM position_snapshots").fetchone()[0]
        by_status_rows = conn.execute("SELECT status, COUNT(*) AS cnt FROM events GROUP BY status ORDER BY cnt DESC").fetchall()
        by_action_rows = conn.execute("SELECT action, COUNT(*) AS cnt FROM events GROUP BY action ORDER BY cnt DESC").fetchall()
        last_event_row = conn.execute("SELECT created_at, action, status FROM events ORDER BY created_at DESC LIMIT 1").fetchone()

    return {
        "generated_at": now_iso(),
        "dry_run": DRY_RUN,
        "journal_db": str(JOURNAL_DB_PATH),
        "reports_dir": str(REPORTS_DIR),
        "total_events": total_events,
        "total_snapshots": total_snapshots,
        "by_status": {row["status"]: row["cnt"] for row in by_status_rows},
        "by_action": {row["action"]: row["cnt"] for row in by_action_rows},
        "last_event": dict(last_event_row) if last_event_row else None,
    }


def _load_json_dict(raw: Any) -> Dict[str, Any]:
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        loaded = json.loads(raw)
        return loaded if isinstance(loaded, dict) else {}
    except Exception:
        return {}


def build_event_report_rows(events: List[Dict[str, Any]], snapshots: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    snapshots_by_event: Dict[str, List[Dict[str, Any]]] = {}
    for snapshot in snapshots:
        snapshots_by_event.setdefault(str(snapshot.get("event_id")), []).append(snapshot)

    rows: List[Dict[str, Any]] = []

    for event in events:
        event_id = str(event.get("event_id"))
        related = snapshots_by_event.get(event_id, [])
        pre_snapshot = next((s for s in related if str(s.get("stage", "")).startswith("pre_")), None)
        post_snapshot = next((s for s in related if str(s.get("stage", "")).startswith("post_")), None)
        rejected_snapshot = next((s for s in related if "rejected" in str(s.get("stage", ""))), None)

        result_json = _load_json_dict(event.get("result_json"))
        request_json = _load_json_dict(event.get("request_json"))
        reconcile = result_json.get("reconcile", {})
        pre_from_result = snapshot_to_report_dict(_load_json_dict(reconcile.get("pre_snapshot")))
        post_from_result = snapshot_to_report_dict(_load_json_dict(reconcile.get("post_snapshot")))

        effective_pre = snapshot_to_report_dict(pre_snapshot) if pre_snapshot else pre_from_result
        effective_post = (
            snapshot_to_report_dict(post_snapshot)
            if post_snapshot
            else snapshot_to_report_dict(rejected_snapshot)
            if rejected_snapshot
            else post_from_result
        )
        effective_post_extra = _load_json_dict(effective_post.get("extra_json"))

        requested_qty = first_number(request_json.get("qtyCoin"), result_json.get("qty"), result_json.get("requested_qty"), allow_zero=True)
        closed_qty = first_number(result_json.get("closed_qty"), allow_zero=True)
        requested_trigger_price = first_number(request_json.get("triggerPrice"))
        requested_fill_price = first_number(request_json.get("fillPrice"))
        requested_entry_price = first_number(request_json.get("entryPrice"))
        requested_theoretical_avg = first_number(request_json.get("theoreticalAvg"))
        requested_tp_price = first_number(request_json.get("tpPrice"), result_json.get("tp_price"))
        requested_lot_tp_price = first_number(request_json.get("lotTpPrice"), result_json.get("lot_tp_price"))
        requested_current_close = first_number(request_json.get("currentClose"))
        pre_entry_price = first_number(effective_pre.get("entry_price"))
        post_entry_price = first_number(effective_post.get("entry_price"))
        pre_mark_price = first_number(effective_pre.get("mark_price"))
        post_mark_price = first_number(effective_post.get("mark_price"))
        market_price = first_number(effective_post.get("market_price"), result_json.get("market_price"), requested_current_close)

        action = str(event.get("action") or "").upper()
        if action in {"FIRST_SHORT", "RESTART_SHORT", "DCA_SHORT"}:
            actual_entry_price = first_number(
                result_json.get("entry_price"),
                effective_post_extra.get("entry_price"),
                post_entry_price,
                requested_fill_price,
                requested_entry_price,
                requested_theoretical_avg,
            )
            actual_exit_price = None
        else:
            actual_entry_price = first_number(
                result_json.get("entry_price"),
                effective_post_extra.get("entry_price"),
                pre_entry_price,
                requested_entry_price,
                requested_theoretical_avg,
            )
            actual_exit_price = first_number(
                result_json.get("exit_price"),
                effective_post_extra.get("exit_price"),
                market_price,
                requested_current_close,
                requested_fill_price,
            )

        row = {
            "created_at": event.get("created_at"),
            "symbol": event.get("symbol_compact") or event.get("symbol_raw"),
            "action": event.get("action"),
            "status": event.get("status"),
            "mode": event.get("mode"),
            "dry_run": event.get("dry_run"),
            "order_id": event.get("order_id"),
            "lot_tag": event.get("lot_tag"),
            "reason": event.get("reason"),
            "pre_side": effective_pre.get("side"),
            "pre_contracts": effective_pre.get("contracts"),
            "pre_entry_price": pre_entry_price,
            "pre_mark_price": pre_mark_price,
            "post_side": effective_post.get("side"),
            "post_contracts": effective_post.get("contracts"),
            "post_entry_price": post_entry_price,
            "post_mark_price": post_mark_price,
            "requested_qty": requested_qty,
            "closed_qty": closed_qty,
            "requested_trigger_price": requested_trigger_price,
            "requested_fill_price": requested_fill_price,
            "requested_entry_price": requested_entry_price,
            "requested_theoretical_avg": requested_theoretical_avg,
            "requested_current_close": requested_current_close,
            "actual_entry_price": actual_entry_price,
            "actual_exit_price": actual_exit_price,
            "market_price": market_price,
            "tp_price": requested_tp_price,
            "lot_tp_price": requested_lot_tp_price,
            "reconcile_status": reconcile.get("reconcile_status") or _load_json_dict(effective_post.get("extra_json")).get("reconcile_status"),
            "error_text": event.get("error_text"),
        }

        if row["reason"] is None:
            row["reason"] = request_json.get("reason")

        rows.append(row)

    return rows


def autosize_worksheet(ws: Any) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def build_excel_report() -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed. Install requirements.txt first.") from exc

    summary = build_summary()
    events = fetch_recent_events(limit=5000)
    snapshots = fetch_recent_snapshots(limit=10000)
    event_report_rows = build_event_report_rows(events, snapshots)

    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "Summary"
    ws_event_report = workbook.create_sheet("Event Report")
    ws_act_event_report = workbook.create_sheet("Act Event Report")
    ws_events = workbook.create_sheet("Events")
    ws_snapshots = workbook.create_sheet("Snapshots")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    success_fill = PatternFill("solid", fgColor="C6EFCE")
    warning_fill = PatternFill("solid", fgColor="FFEB9C")
    error_fill = PatternFill("solid", fgColor="FFC7CE")
    neutral_fill = PatternFill("solid", fgColor="D9EAD3")

    summary_rows = [
        ("Generated At", summary["generated_at"]),
        ("Dry Run", summary["dry_run"]),
        ("Journal DB", summary["journal_db"]),
        ("Reports Dir", summary["reports_dir"]),
        ("Total Events", summary["total_events"]),
        ("Total Snapshots", summary["total_snapshots"]),
        ("Last Event", safe_json(summary["last_event"]) if summary["last_event"] else ""),
        ("By Status", safe_json(summary["by_status"])),
        ("By Action", safe_json(summary["by_action"])),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=row_index, column=1, value=label)
        ws_summary.cell(row=row_index, column=2, value=value)

    event_report_headers = [
        "created_at",
        "symbol",
        "action",
        "status",
        "mode",
        "dry_run",
        "order_id",
        "lot_tag",
        "reason",
        "requested_qty",
        "closed_qty",
        "pre_side",
        "pre_contracts",
        "pre_entry_price",
        "pre_mark_price",
        "post_side",
        "post_contracts",
        "post_entry_price",
        "post_mark_price",
        "requested_trigger_price",
        "requested_fill_price",
        "requested_entry_price",
        "requested_theoretical_avg",
        "requested_current_close",
        "actual_entry_price",
        "actual_exit_price",
        "market_price",
        "tp_price",
        "lot_tp_price",
        "reconcile_status",
        "error_text",
    ]
    for col, header in enumerate(event_report_headers, start=1):
        cell = ws_event_report.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(event_report_rows, start=2):
        for col, header in enumerate(event_report_headers, start=1):
            ws_event_report.cell(row=row_idx, column=col, value=row.get(header))

        status_value = str(row.get("status") or "").lower()
        fill = None
        if status_value == "executed":
            fill = success_fill
        elif status_value in {"tp_not_reached", "duplicate_ignored", "startup_synced", "startup_skipped", "startup_partial"}:
            fill = warning_fill
        elif status_value in {"error", "startup_error"}:
            fill = error_fill
        elif status_value in {"no_position", "ignored"}:
            fill = neutral_fill

        if fill is not None:
            for col in range(1, len(event_report_headers) + 1):
                ws_event_report.cell(row=row_idx, column=col).fill = fill

    act_event_report_headers = [
        "created_at",
        "symbol",
        "action",
        "status",
        "mode",
        "dry_run",
        "order_id",
        "lot_tag",
        "reason",
        "requested_qty",
        "closed_qty",
        "requested_trigger_price",
        "requested_fill_price",
        "requested_entry_price",
        "requested_theoretical_avg",
        "requested_current_close",
        "actual_entry_price",
        "actual_exit_price",
        "tp_price",
        "lot_tp_price",
        "market_price",
        "pre_side",
        "pre_contracts",
        "pre_entry_price",
        "pre_mark_price",
        "post_side",
        "post_contracts",
        "post_entry_price",
        "post_mark_price",
        "reconcile_status",
        "error_text",
    ]
    for col, header in enumerate(act_event_report_headers, start=1):
        cell = ws_act_event_report.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(event_report_rows, start=2):
        for col, header in enumerate(act_event_report_headers, start=1):
            ws_act_event_report.cell(row=row_idx, column=col, value=row.get(header))

        status_value = str(row.get("status") or "").lower()
        fill = None
        if status_value == "executed":
            fill = success_fill
        elif status_value in {"tp_not_reached", "duplicate_ignored", "startup_synced", "startup_skipped", "startup_partial"}:
            fill = warning_fill
        elif status_value in {"error", "startup_error"}:
            fill = error_fill
        elif status_value in {"no_position", "ignored"}:
            fill = neutral_fill

        if fill is not None:
            for col in range(1, len(act_event_report_headers) + 1):
                ws_act_event_report.cell(row=row_idx, column=col).fill = fill

    event_headers = [
        "event_id",
        "created_at",
        "updated_at",
        "action",
        "status",
        "source",
        "mode",
        "dry_run",
        "symbol_raw",
        "symbol_compact",
        "ccxt_symbol",
        "order_id",
        "lot_tag",
        "reason",
        "error_text",
        "request_json",
        "result_json",
    ]
    for col, header in enumerate(event_headers, start=1):
        cell = ws_events.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(events, start=2):
        for col, header in enumerate(event_headers, start=1):
            ws_events.cell(row=row_idx, column=col, value=row.get(header))

    snapshot_headers = [
        "id",
        "event_id",
        "created_at",
        "stage",
        "symbol_compact",
        "ccxt_symbol",
        "side",
        "contracts",
        "entry_price",
        "mark_price",
        "unrealized_pnl",
        "market_price",
        "order_id",
        "extra_json",
    ]
    for col, header in enumerate(snapshot_headers, start=1):
        cell = ws_snapshots.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(snapshots, start=2):
        for col, header in enumerate(snapshot_headers, start=1):
            ws_snapshots.cell(row=row_idx, column=col, value=row.get(header))

    autosize_worksheet(ws_summary)
    autosize_worksheet(ws_event_report)
    autosize_worksheet(ws_act_event_report)
    autosize_worksheet(ws_events)
    autosize_worksheet(ws_snapshots)

    report_path = REPORTS_DIR / REPORT_FILENAME
    workbook.save(report_path)
    return report_path


@app.post("/webhook")
async def webhook(request: Request, background_tasks: BackgroundTasks) -> Dict[str, Any]:
    try:
        body = await request.json()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid JSON body: {exc}") from exc

    try:
        payload = WebhookPayload(**body)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Invalid payload: {exc}") from exc

    verify_secret(payload, request)

    if not register_event_once(payload):
        logger.warning("Duplicate webhook ignored | action=%s | symbol=%s | orderId=%s", payload.action, payload.symbol, payload.orderId)
        send_telegram_log(build_duplicate_alert_message(payload))
        return {
            "status": "duplicate_ignored",
            "action": payload.action,
            "symbol": payload.symbol,
            "server_time": now_iso(),
        }

    event_id = new_event_id()
    insert_event_shell(event_id, payload)

    logger.info(
        "Webhook accepted | event_id=%s | action=%s | symbol=%s | orderId=%s | qtyCoin=%s",
        event_id,
        payload.action,
        payload.symbol,
        payload.orderId,
        payload.qtyCoin,
    )
    send_telegram_log(build_alert_accept_message(event_id, payload))

    background_tasks.add_task(background_process, payload, event_id)
    return {
        "status": "accepted",
        "event_id": event_id,
        "action": payload.action,
        "symbol": payload.symbol,
        "server_time": now_iso(),
        "dry_run": DRY_RUN,
    }


@app.get("/health")
def health() -> Dict[str, Any]:
    return {
        "status": "ok",
        "time": now_iso(),
        "exchange": "BingX",
        "markets_loaded": _markets_loaded,
        "dry_run": DRY_RUN,
        "allowed_symbols": sorted(ALLOWED_SYMBOLS),
        "default_margin_mode": DEFAULT_MARGIN_MODE,
        "journal_db": str(JOURNAL_DB_PATH),
        "reports_dir": str(REPORTS_DIR),
    }


@app.get("/events")
def list_events(limit: int = 100) -> Dict[str, Any]:
    limit = max(1, min(limit, 1000))
    return {"items": fetch_recent_events(limit)}


@app.get("/snapshots")
def list_snapshots(limit: int = 200) -> Dict[str, Any]:
    limit = max(1, min(limit, 2000))
    return {"items": fetch_recent_snapshots(limit)}


@app.get("/report/summary")
def report_summary() -> Dict[str, Any]:
    return build_summary()


@app.get("/report/excel")
def report_excel() -> FileResponse:
    report_path = build_excel_report()
    return FileResponse(
        path=report_path,
        filename=report_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app_hybrid_reporting:app", host="0.0.0.0", port=PORT, reload=False)
